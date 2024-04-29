from openpyxl import load_workbook
from os import mkdir 
from os import chdir
from os.path import isdir
import segno

OUT_QRCODES_FILES='qr-codes' # directorio que será armazenado os códigos QR

def main():
  wb = load_workbook('format-doc.xlsx') # arquivo onde estão os nomes de utilizadores
  ws = wb.active

  user_data = ""
  filename = ""

  count = 0
  y = 1

  if not isdir(OUT_QRCODES_FILES):
    mkdir(OUT_QRCODES_FILES)
  chdir(OUT_QRCODES_FILES)

  try:
    for row in ws:
      for value in row:
        count+=1
        if count > 4:
          count = 1
          user_data = ""
          y += 1
          username = ws.cell(row=y, column=2).value
          print()
          filename = username.strip().lower().replace(' ', '_')

        column_name = ws.cell(row=1, column=count).value
        cell_value = ws.cell(row=y, column=count).value
        user_data += f'{column_name}: {cell_value}\n'
      user_data += """
        site: 0xff.com\n
        email: 0xff@local.dev\n
        2024 @ 0xff.
      """
      qrcode = segno.make(user_data)
      if filename:
        qrcode.save(
            f'qr-{filename}.png', 
            scale=10, 
            kind="png"
          )
        print(f'file to save: qrcode-{filename}.png')
  except Exception as err: 
    print(f'Total users: {y-2}', err)
     
if __name__=="__main__":
  main()