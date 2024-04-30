from openpyxl import load_workbook
import segno
from os.path import isdir
from os import mkdir 
from os import chdir 

def main():
  try:
    wb = load_workbook('format-doc.xlsx')
    
    CONFIG_SHEET = {
      "max_cols": 4, # ajust the number of coluns
      "max_rows": 3, # ajust the numbers of rows including the row headers
    }

    DIR_OUT = 'qr-codes/' # in local directory of project

    if not isdir(DIR_OUT):
      mkdir(DIR_OUT)

    chdir(DIR_OUT)

    TOTAL_LIST = 0
    FILE_OUT = ''
    QR_DATA = ''

    hackerman = wb['Sheet'] # select the sheet where contain the data

    iter_sheet_data = enumerate(hackerman.iter_rows(
      max_col=CONFIG_SHEET['max_cols'], 
      max_row=CONFIG_SHEET['max_rows'], 
      values_only=True
    ))

    cols = ()

    for index, row in iter_sheet_data:
      if index == 0:
        cols = row
        continue
      TOTAL_LIST += 1
      QR_DATA = ''
      for index_col in range(len(cols)):
        #print(f"{cols[index_col]} : {row[index_col]}")
        QR_DATA += f"{cols[index_col]}: {row[index_col]}\n"

        if cols[index_col] == 'NOME COMPLETO':
          FILE_OUT = row[index_col].lower().split(' ')[0] + '.png'
      if not FILE_OUT:
        raise Exception("your document not have a NOME COMPLETO column")
      
      QR_DATA += f'SITE: hello.com\nEMAIL: hello@local.lan' # signature
      
      qrcode = segno.make(QR_DATA)
     
      qrcode.save(
          FILE_OUT, 
          scale=10, 
          kind="png"
      )
      
      print(f'file: {FILE_OUT}')

      print()
    print(f'total: {TOTAL_LIST} file generated')
  except Exception as err:
    print(err)

if __name__ == '__main__':
  main()

